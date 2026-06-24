# workout_generator.py
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

def generate_excel(program_data, filepath):
    """
    program_data: {"weeks": [{"week_number": 1, "days": [...]}, ...]}
    """
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, size=14)
    day_font = Font(bold=True, size=12, color="FFFFFF")
    day_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    ex_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for week in program_data["weeks"]:
        ws = wb.create_sheet(title=f"Неделя {week['week_number']}")
        # Заголовок листа — объединяем 4 столбца (A-D)
        ws.merge_cells('A1:D1')
        ws['A1'] = f"Программа тренировок – Неделя {week['week_number']}"
        ws['A1'].font = header_font
        ws['A1'].alignment = Alignment(horizontal='center')

        row = 3
        for day in week["days"]:
            # Название дня — объединяем 4 столбца
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            cell = ws.cell(row=row, column=1, value=day["day_name"])
            cell.font = day_font
            cell.fill = day_fill
            cell.alignment = Alignment(horizontal='center')
            row += 1

            # Заголовки таблицы (4 столбца)
            headers = ["Упражнение", "Подходы", "Повторения", "Примечания"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            row += 1

            # Данные упражнений
            for ex in day["exercises"]:
                ws.cell(row=row, column=1, value=ex["exercise"]).border = thin_border
                ws.cell(row=row, column=2, value=ex["sets"]).border = thin_border
                ws.cell(row=row, column=3, value=ex["reps"]).border = thin_border
                ws.cell(row=row, column=4, value=ex.get("notes", "")).border = thin_border
                # выравнивание и перенос текста
                for c in range(1, 5):
                    ws.cell(row=row, column=c).alignment = Alignment(vertical='center', wrap_text=True)
                row += 1
            row += 2  # отступ между днями

        # Автоподбор ширины столбцов (чуть увеличены для комфорта)
        col_widths = [38, 12, 16, 32]  # Упражнение, Подходы, Повторения, Примечания
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(filepath)